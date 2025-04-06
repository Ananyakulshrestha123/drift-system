import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load Excel file
jira_file = "JiraSheet.xlsx"
data = pd.read_excel(jira_file)

# Load keywords from config
with open("config.json", "r") as f:
    config = json.load(f)

dev_keywords = config["development_keywords"]
test_keywords = config["testing_keywords"]

# Function to categorize tasks
def categorize_task(description, criteria):
    text = (str(description).lower() if pd.notna(description) else "") + " " + \
           (str(criteria).lower() if pd.notna(criteria) else "")
    if any(word in text for word in dev_keywords):
        return "Development"
    elif any(word in text for word in test_keywords):
        return "Testing"
    else:
        return "Other"

# Apply categorization
data["Category"] = data.apply(lambda row: categorize_task(row["Description"], row["Acceptance Criteria"]), axis=1)

# Group and summarize
summary = data.groupby(["Assignee", "Category"]).agg({"Story Points": "sum"}).unstack(fill_value=0).reset_index()
summary.columns = ['Assignee'] + list(summary.columns.get_level_values(1)[1:])
summary["Total Story Points"] = summary[["Development", "Testing", "Other"]].sum(axis=1)

# Save everything to Excel
output_file = "Jira_Report_Final.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    summary.to_excel(writer, sheet_name="Story Point Summary", index=False)

    # Sheets for blank fields
    blank_fields = ["Description", "Acceptance Criteria", "Fix version", "Affected version", "Epic Link"]
    required_cols = ["Assignee", "Epic Link", "Status", "Key", "Issue Type"]
    
    for field in blank_fields:
        if field in data.columns:
            blank_data = data[data[field].isna()][required_cols]
            if not blank_data.empty:
                sheet_name = f"Blank {field}"
                blank_data.to_excel(writer, sheet_name=sheet_name, index=False)

# Highlight total > 8
wb = load_workbook(output_file)
ws = wb["Story Point Summary"]
highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in ws.iter_rows(min_row=2, min_col=ws.max_column, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None and cell.value > 8:
            cell.fill = highlight_fill

wb.save(output_file)
