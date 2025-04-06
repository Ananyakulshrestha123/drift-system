import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load Excel file
jira_file = "JiraSheet_1.xlsx"
try:
    data = pd.read_excel(jira_file)
except FileNotFoundError:
    print(f"❌ File not found: {jira_file}")
    exit(1)

# Load keywords from config.json
with open("config.json", "r") as f:
    config = json.load(f)

dev_keywords = config.get("development_keywords", [])
test_keywords = config.get("testing_keywords", [])

# Function to categorize tasks
def categorize_task(description, criteria):
    text = (str(description).lower() if pd.notna(description) else "") + " " or \
           (str(criteria).lower() if pd.notna(criteria) else "")
    if any(word in text for word in dev_keywords):
        return "Development"
    elif any(word in text for word in test_keywords):
        return "Testing"
    else:
        return "Other"

# Apply category
data["Category"] = data.apply(lambda row: categorize_task(row.get("Description"), row.get("Acceptance Criteria")), axis=1)

# Detect Sprint Type
def get_sprint_group(sprint_value):
    if pd.isna(sprint_value):
        return "Unknown"
    sprint_str = str(sprint_value).lower()
    if "fastfinder" in sprint_str:
        return "FastFinder"
    elif "fastbreak" in sprint_str:
        return "FastBreak"
    else:
        return "Other"

data["Sprint Group"] = data["Sprint"].apply(get_sprint_group)

# Create output Excel
output_file = "Jira_Report_Final_With_Sprint_Grouping.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    required_cols = ["Assignee", "Epic Link", "Status", "Key", "Issue Type"]

    for sprint_group in data["Sprint Group"].unique():
        sprint_data = data[data["Sprint Group"] == sprint_group]
        if sprint_data.empty:
            continue

        # ✅ Updated: Flat Story Summary Table (like format.xlsx)
        flat_columns = [
            "Assignee", "Epic Link", "Status", "Key", "Issue Type", "Category",
            "Sprint", "Fix version", "Affected version", "Story Points", "Reporter"
        ]
        available_columns = [col for col in flat_columns if col in sprint_data.columns]
        flat_summary = sprint_data[available_columns].copy()

        # Add total row
        total_story_points = flat_summary["Story Points"].sum() if "Story Points" in flat_summary else 0
        total_row = {col: "" for col in available_columns}
        total_row["Assignee"] = "Total"
        total_row["Story Points"] = total_story_points
        flat_summary = pd.concat([flat_summary, pd.DataFrame([total_row])], ignore_index=True)

        flat_summary.to_excel(writer, sheet_name=f"{sprint_group} Story Summary", index=False)

        # 🧠 Assignee-Category summary
        summary = sprint_data.groupby(["Assignee", "Category"])["Story Points"].sum().unstack(fill_value=0).reset_index()
        summary.columns.name = None

        # 🛡️ Safely add missing columns
        for col in ["Development", "Testing", "Other"]:
            if col not in summary.columns:
                summary[col] = 0

        summary["Total Story Points"] = summary["Development"] + summary["Testing"] + summary["Other"]
        total_row = pd.DataFrame([{
            "Assignee": "Total",
            "Development": summary["Development"].sum(),
            "Testing": summary["Testing"].sum(),
            "Other": summary["Other"].sum(),
            "Total Story Points": summary["Total Story Points"].sum()
        }])
        summary = pd.concat([summary, total_row], ignore_index=True)

        summary.to_excel(writer, sheet_name=f"{sprint_group} Summary", index=False)

        # 📊 Sprint Group - Category Summary
        category_summary = sprint_data.groupby("Category")["Story Points"].sum().reset_index()
        category_summary.columns = ["Category", "Total Story Points"]
        total_row = pd.DataFrame([["Total", category_summary["Total Story Points"].sum()]], columns=category_summary.columns)
        category_summary = pd.concat([category_summary, total_row], ignore_index=True)
        category_summary.to_excel(writer, sheet_name=f"{sprint_group} Category Summary", index=False)

    # 🔍 Blank fields check
    blank_fields = ["Description", "Acceptance Criteria", "Fix version", "Affected version", "Epic Link"]
    for field in blank_fields:
        if field in data.columns:
            blank_data = data[data[field].isna()]
            if not blank_data.empty:
                blank_data_subset = blank_data[required_cols + [field]]
                blank_data_subset.to_excel(writer, sheet_name=f"Blank {field}", index=False)

    # 🧵 Unique Epic Link sheet
    unique_epics = data.dropna(subset=["Epic Link"]).drop_duplicates(subset=["Epic Link"])
    unique_epics[["Epic Link"]].to_excel(writer, sheet_name="Unique Epics", index=False)

# 🎨 Highlight Total Story Points > 8 in summary sheets
wb = load_workbook(output_file)
highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for sheet_name in wb.sheetnames:
    if "Summary" in sheet_name and "Category" not in sheet_name and "Story" not in sheet_name:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            total_cell = row[ws.max_column - 1]
            if total_cell.value and isinstance(total_cell.value, (int, float)) and total_cell.value > 8:
                total_cell.fill = highlight_fill

wb.save(output_file)
print(f"✅ Final report generated with flat story summaries: {output_file}")
