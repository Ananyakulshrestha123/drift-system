import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load Excel file
jira_file = "JiraSheet_1.xlsx"
data = pd.read_excel(jira_file)

# Load keywords from config.json
with open("config.json", "r") as f:
    config = json.load(f)

dev_keywords = config["development_keywords"]
test_keywords = config["testing_keywords"]

# Function to categorize tasks
def categorize_task(description, criteria):
    text = (str(description).lower() if pd.notna(description) else "") + " " + \
           (str(criteria).lower() if pd.notna(criteria) else "")
    if any(word in text for word in dev_keywords):
        return "Development"
    elif any(word in text for word in test_keywords):
        return "Testing"
    else:
        return "Other"

# Categorize each row
data["Category"] = data.apply(lambda row: categorize_task(row["Description"], row["Acceptance Criteria"]), axis=1)

# Summarize story points by Assignee and Category
summary = data.groupby(["Assignee", "Category"]).agg({"Story Points": "sum"}).unstack(fill_value=0).reset_index()
summary.columns = ['Assignee'] + list(summary.columns.get_level_values(1)[1:])
summary["Total Story Points"] = summary[["Development", "Testing", "Other"]].sum(axis=1)

# Write to Excel
output_file = "Jira_Report_Final_v4.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Sheet 1: Story point summary
    summary.to_excel(writer, sheet_name="Story Point Summary", index=False)

    # Sheets 2-N: Rows with blank values in each key field + include the blank field column itself
    blank_fields = ["Description", "Acceptance Criteria", "Fix Version", "Affected Version", "Epic Link"]
    required_cols = ["Assignee", "Epic Link", "Status", "Key", "Issue Type"]
    
    for field in blank_fields:
        if field in data.columns:
            blank_data = data[data[field].isna()][required_cols + [field]]
            if not blank_data.empty:
                blank_data.to_excel(writer, sheet_name=f"Blank {field}", index=False)

    # Last Sheet: Unique epic links with no blank fields
    complete_epics = data.dropna(subset=blank_fields)
    unique_epics = complete_epics.drop_duplicates(subset=["Epic Link"])
    unique_epics[["Epic Link"]].to_excel(writer, sheet_name="Unique Epics", index=False)

# Highlight users with Total Story Points > 8
wb = load_workbook(output_file)
ws = wb["Story Point Summary"]
highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in ws.iter_rows(min_row=2, min_col=ws.max_column, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None and cell.value > 8:
            cell.fill = highlight_fill

wb.save(output_file)
print(f"✅ Report generated: {output_file}")
