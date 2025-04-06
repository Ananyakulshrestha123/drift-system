import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Load JIRA Excel File ===
jira_file = "JiraSheet.xlsx"
data = pd.read_excel(jira_file)

# === Load Config (Keywords) ===
with open("config.json", "r") as f:
    config = json.load(f)

dev_keywords = config["development_keywords"]
test_keywords = config["testing_keywords"]

# === Categorize Tasks ===
def categorize_task(description, criteria):
    text = (str(description).lower() if pd.notna(description) else "") or \
           (str(criteria).lower() if pd.notna(criteria) else "")
    if any(word in text for word in dev_keywords):
        return "Development"
    elif any(word in text for word in test_keywords):
        return "Testing"
    else:
        return "Other"

data["Category"] = data.apply(
    lambda row: categorize_task(row.get("Description", ""), row.get("Acceptance Criteria", "")),
    axis=1
)

# === Summary Table ===
summary = data.groupby(["Assignee", "Category"]).agg({"Story Points": "sum"}).unstack(fill_value=0).reset_index()

# Flatten MultiIndex columns
columns = ['Assignee'] + list(summary.columns.get_level_values(1)[1:])
summary.columns = columns

# Ensure required columns
for col in ["Development", "Testing", "Other"]:
    if col not in summary.columns:
        summary[col] = 0

# Add Total Story Points
summary["Total Story Points"] = summary[["Development", "Testing"]].sum(axis=1)

# Save to Excel
summary_output_file = "Timesheet_Summary_Output.xlsx"
summary.to_excel(summary_output_file, index=False)

# === Highlight "Total Story Points" if Dev + Test > 8 ===
wb = load_workbook(summary_output_file)
ws = wb.active

# Red fill
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Get column indexes
header = [cell.value for cell in ws[1]]
dev_col = header.index("Development") + 1 if "Development" in header else None
test_col = header.index("Testing") + 1 if "Testing" in header else None
total_col = header.index("Total Story Points") + 1

# Highlight rows
for row in range(2, ws.max_row + 1):
    dev = ws.cell(row=row, column=dev_col).value or 0
    test = ws.cell(row=row, column=test_col).value or 0
    try:
        if float(dev) + float(test) > 8:
            ws.cell(row=row, column=total_col).fill = red_fill
    except:
        continue

wb.save(summary_output_file)
print(f"✅ Timesheet saved to {summary_output_file} with highlights.")

# === Blank Field Reports ===
blank_fields_to_check = {
    "Description": [],
    "Acceptance Criteria": [],
    "Fix Version": [],
    "Affects Version/s": [],
    "Epic Link": []
}

required_columns = ["Assignee", "Epic Link", "Status", "Key", "Issue Type"]

# Ensure required columns exist
for col in required_columns:
    if col not in data.columns:
        data[col] = None

# Find and collect blanks
for field in blank_fields_to_check:
    if field not in data.columns:
        data[field] = None
    blanks = data[data[field].isna()]
    if not blanks.empty:
        blank_fields_to_check[field] = blanks[required_columns]

# Save blank reports to Excel
blank_output_file = "JIRA_Blank_Fields_Summary.xlsx"
with pd.ExcelWriter(blank_output_file, engine='openpyxl') as writer:
    for field, df in blank_fields_to_check.items():
        if not df.empty:
            sheet_name = f"Blank_{field[:28]}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"✅ Blank fields report saved to {blank_output_file}")
